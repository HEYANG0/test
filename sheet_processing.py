import time

import openpyxl
from re import findall, search, sub


def find_date(date):
    par = r'\w+[0-9]+[-]\w+[0-9]+'
    fdate = findall(par, date)
    return fdate


def resolve_date(f_date):
    prefix = search('[a-zA-Z]+', f_date)
    date = sub('[a-zA-Z]', '', f_date)  # 替换字母为空
    date = date.split('-')
    date = [int(x) for x in date]
    if date[0] > date[1]:
        date[0], date[1] = date[1], date[0]
    date_1 = date[0]  # 初始化date_1
    while date_1 < (date[1]):
        date_1 += 1
        date.append(date_1)
    date.remove(date[1])
    date = [prefix.group() + str(x) for x in date]
    date = ','.join(date)
    return date


def replace_date(f_date, date1):
    for i in range(len(f_date)):
        date = resolve_date(f_date[i])
        date1 = date1.replace(f_date[i], date)
    return date1


def r_date(date1):
    fdate = find_date(date1)
    replace_date1 = sub('，', ',', replace_date(fdate, date1))
    replace_date1 = sub(' ', '', replace_date1)
    return replace_date1


def tag_process(file_name, column):
    wb = openpyxl.load_workbook(filename=file_name)
    sheet_name = wb.sheetnames
    ws = wb[sheet_name[0]]
    # 在后面插入
    ws.insert_cols(5, 2)
    next_column = chr(ord(column) + 1)  # 新增的第一列的列序号
    nnext_column = chr(ord(column) + 2)  # 新增的第二列的列序号
    for row in ws[column]:
        # print(row.value)
        if (str(row.value)).isspace() or row.value == None:
            ws[next_column + f'{row.row}'] = ''
            ws[nnext_column + f'{row.row}'] = ''
        else:
            date = r_date(str(row.value))
            ws[next_column + f'{row.row}'] = date
            num = date.count(',')
            # print(num)
            ws[nnext_column + f'{row.row}'] = int(num) + 1
        # print(ws['j1'].value)
        # date.append(r_date(str(row.value)))
    # for row1 in ws['H']:
    #     row1.value = date[(row1.row - 1)]
    # print(row1.value)
    # print(date)
    ws[next_column+'1'] = '处理后的位号'
    ws[nnext_column+'1'] = '位号数量'
    wb.save(file_name)  # 写入文件
    print("文件位号处理完毕，窗口将在3秒后关闭")
    time.sleep(3)


# global filename
# global place_column


def n_input():
    global filename
    global place_column
    filename = input("请直接拖入xlsx文件:")
    filename = filename.replace("& ", "").replace("\'", "").replace("\"", "")
    place_column = input("请输入位号的列序号(比如A列):")
    return filename, place_column


n_input()

tag_process(filename, column=place_column)

# tag_processing("H2", "H79")
